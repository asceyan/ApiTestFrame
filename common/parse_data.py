# -*- coding: utf-8 -*-
# @ Time: 2021/2/22 14:26
# @ Author: YanZhiJia
# @ Email: asce_yan@foxmail.com
# @ File: parse_data
# @ Description: 读取文件并转换成数据格式的方法
import os
import yaml
from string import Template
from openpyxl import load_workbook

BASE_PATH = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
DATA_PATH = os.path.join(BASE_PATH, 'test')


class ParseExcel(object):
    """读取excel格式数据"""
    def __init__(self, excel_path, sheet):
        print(excel_path, sheet)
        self.wb = load_workbook(excel_path)
        self.sheet = self.wb[sheet]
        self.maxRowNum = self.sheet.max_row

    def get_data_from_sheet(self):
        data_list = []
        for line in self.sheet.rows:
            tmpList=[]
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            data_list.append(tmpList)
        return data_list[2:]


class ReadYaml:
    @staticmethod
    def read_yaml(yamlpath, replace=False, **kwargs):
        """
        读取yaml文件内容，设置参数replace=True可通过Template方法，替换yaml文件中变量
        """
        if not os.path.isfile(yamlpath):
            raise FileNotFoundError("文件路径不存在，请检查文件路径是否正确 %s" % yamlpath)
        if replace:
            with open(yamlpath, encoding="utf-8") as f:
                re = Template(f.read()).safe_substitute(**kwargs)
                dict_data = yaml.safe_load(re)
            return dict_data
        else:
            with open(yamlpath, encoding="utf-8") as f:
                dict_data = yaml.safe_load(f.read())
            return dict_data


class ParseCSV(object):
    """读取csv文件数据，并组合为列表+字典格式"""
    def __init__(self, filename):
        try:
            with open(filename, encoding='utf-8') as file:
                self.content = file.readlines()    # 按行读取并放入到一个列表中
        except:
            raise FileNotFoundError("文件路径不存在，请检查文件路径是否正确 %s" % filename)

    def read_csv(self):
        data_list = []
        try:
            key_list = self.content[0].strip().split(',')    # 把第1行split为字典的key值
            for i in range(1, len(self.content)):
                if self.content[i].strip() == '':        # 判断列表数据是否存在空行
                    continue
                if self.content[i].startswith('#'):      # 省略开头为#的数据
                    continue
                line_list = self.content[i].strip().split(',')
                if len(line_list) < len(key_list):
                    continue
                data_dict = {}
                for j in range(len(key_list)):
                    data_dict[key_list[j]] = line_list[j]    # 用key_list里面的条目作为字典Key，用第2行开始的每1行的条目作为字典的Value
                data_list.append(data_dict)
            return data_list
        except:
            raise Exception(f"The data format is incorrect. Please reenter it")


if __name__ == '__main__':
    path = os.path.join(DATA_PATH, 'user.csv')
    data = ParseCSV(path)
    print(data.read_csv())